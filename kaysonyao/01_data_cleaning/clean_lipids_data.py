"""
clean_lipids_data.py
====================
Converts the Sadovsky untargeted plasma lipidomics Excel workbook into
wide-format CSVs compatible with the DP3 binary/multilabel pipeline.

Input
-----
  Excel workbook: data/lipids/072925 Sadovsky Plasma Lipids Untargeted ALL.xlsx
    Sheet "Plasma POS Lipids"  — positive-mode features, all three batches integrated
    Sheet "Plasma NEG Lipids"  — negative-mode features, all three batches integrated
    Sheet "041625_Sadovsky_Plasma_3Sets_Po"  — raw LipidSearch output for all 3 sets
      processed together (same 3915 rows/lipids as main sheet, same order). Contains
      the Rej quality-control flag used to filter the main sheet.
    Sheet "Sadovsky_Plasma_3Sets_Neg"        — same as above for negative mode.

  The "3Sets" sheets are NOT a separate batch of new samples — they are the raw
  instrument QC output from analysing all three sample batches (051223, 110123,
  041625) together. The main "Plasma POS/NEG Lipids" sheets are the organised
  version of this same data (multi-row header, DP3 sample IDs added) but lack the
  Rej column. The Rej flags from the raw sheets are therefore the authoritative
  quality filter for the main sheets.

  Metadata: data/dp3 master table v2.xlsx
    Sheet "Sheet1": columns used — ID (col 0), gest age del (col 5),
                    group (col 7), subgroup (col 8), omics set# / batch (col 16)

Output
------
  data/cleaned/lipids/normalized_full_results/
    lipids_plasma_cleaned_with_metadata.csv   — all timepoints merged (full matrix)
    lipids_plasma_dropped_missingness_report.csv  — features dropped for missingness
  data/cleaned/lipids/normalized_sliced_by_suffix/
    lipids_plasma_formatted_suffix_A.csv      — timepoint A only
    lipids_plasma_formatted_suffix_B.csv      — timepoint B only
    lipids_plasma_formatted_suffix_C.csv      — timepoint C only
    lipids_plasma_formatted_suffix_D.csv      — timepoint D only
    lipids_plasma_formatted_suffix_E.csv      — timepoint E only

Output CSV format (identical to proteomics/metabolomics pipeline format)
------
  - Index: SampleID (DP3-XXXXZ, where Z is the timepoint letter A–E)
  - Columns: SubjectID, Batch, Group, Subgroup, GestAgeDelivery, <lipid_feature_1>, …
  - Values: log2(peak_area), ComBat-corrected, missing imputed with half-minimum
  - Feature names: prefixed with POS__ or NEG__ to avoid collisions across polarities
    e.g.  POS__TG(50:1)+NH4   NEG__PI(34:2)-H

Pipeline
--------
  1.  Read Rej flags from raw 3Sets sheets — build QC rejection filter.
  2.  Read POS processed sheet → raw peak area matrix (samples × features).
  3.  Read NEG processed sheet → raw peak area matrix.
  4.  Merge POS + NEG (outer join on SampleID).
  5.  Log2(area + 1) transformation.
  6.  Load metadata: Group, Subgroup, GestAgeDelivery, Batch (omics set#).
  7.  Drop samples with missing group or batch labels (logged as warnings).
  8.  ComBat batch correction (wide matrix) — preserves original missingness.
  9.  Missingness filter: drop features with ≥ 20% missing values.
      For dropped features, run Fisher's exact test (group-wise missingness)
      with Benjamini-Hochberg correction; save report CSV.
  10. Half-minimum imputation per feature (log2-space: fill = min_log2 − 1).
  11. Prepend SubjectID, Batch, Group, Subgroup, GestAgeDelivery columns.
  12. Save full matrix CSV.
  13. Slice by timepoint suffix (A–E) and save per-timepoint CSVs.

Pipeline decisions
------------------
  1. ISTD rows (LipidGroup contains "ISTD") are dropped — they are instrument
     internal standards, not biological analytes.
  2. Rejected features: a LipidID is dropped if it is flagged Rej=True in ALL its
     appearances in the raw 3Sets sheet (never Rej=False). LipidIDs that appear
     as non-rejected at least once are kept. The 3Sets raw sheet rows are in the
     same order as the main processed sheet, and the Rej flags map 1-to-1.
  3. Pooled QC columns (Sample ID = "Pooled Cntrl" or "Pooled Control") are dropped.
  4. Duplicate sample columns (same canonical DP3 SampleID after normalisation):
     two sources — (a) same sample re-injected across batches (e.g. DP3-0008C in
     both 041625 and 110123 batches), and (b) EA/EB/EC/ED multi-injection replicates
     that all collapse to the same canonical E timepoint.  In both cases the raw peak
     areas are AVERAGED across all matching columns before log2 transformation, so no
     systematic inflation occurs for samples measured 2–5 times.
  5. Log2 transformation: log2(area + 1) to handle any residual zeros.
  6. Imputation: remaining missing values → half-minimum per analyte in log2 space.
     Since values are already log2-transformed, halving the original abundance equals
     subtracting 1:  log2(x / 2) = log2(x) - 1.
     Fill value = min_log2_observed - 1   (NOT min_log2 / 2, which would be wrong).
  7. Duplicate lipid rows (same LipidID appearing multiple times in the feature rows):
     peak areas are included in the per-sample accumulator and averaged together with
     any other measurements for that (sample, feature) pair.
  8. Non-standard SampleIDs (e.g. "DP3-0024 D", "DP3-0258EA"):
     - Whitespace between subject and timepoint is stripped: "DP3-0024 D" → "DP3-0024D"
     - Multi-character suffixes (e.g. "EA", "EB") → only the first character is kept
       as the timepoint: "DP3-0258EA" → "DP3-0258E"  (A/B/C/D suffix = batch replicate)
     Samples whose timepoint cannot be resolved to A–E are flagged and dropped.

Usage
-----
  # From project root
  python 01_data_cleaning/clean_lipids_data.py

  # Explicit paths
  python 01_data_cleaning/clean_lipids_data.py \\
    --excel  "data/lipids/072925 Sadovsky Plasma Lipids Untargeted ALL.xlsx" \\
    --metadata  "data/dp3 master table v2.xlsx" \\
    --output-full-dir   data/cleaned/lipids/normalized_full_results \\
    --output-sliced-dir data/cleaned/lipids/normalized_sliced_by_suffix
"""

import argparse
import logging
import os
import re
import sys
from collections import defaultdict

import numpy as np
import pandas as pd

# Make utilities.py importable whether the script is run from the project root
# (python 01_data_cleaning/clean_lipids_data.py) or from its own directory.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utilities import (
    CUTOFF_PERCENT_MISSING,
    combat_normalize_wide,
    half_min_impute_wide,
    missingness_filter_and_group_check,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# PCA helpers (mirrors proteomics_diagnostics.py)
# ---------------------------------------------------------------------------

def _generate_pca_plots(
    X_pre: pd.DataFrame,
    X_post: pd.DataFrame,
    batch_labels: pd.Series,
    group_labels: pd.Series,
    out_dir: str,
) -> None:
    """
    Save side-by-side pre/post ComBat PCA plots colored by batch and by group.
    Matches the pattern used in proteomics_diagnostics.plot_pca_pre_post_comparison.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from sklearn.decomposition import PCA as _PCA
    except ImportError:
        logger.warning("matplotlib or scikit-learn not installed — PCA plots skipped.")
        return

    os.makedirs(out_dir, exist_ok=True)

    def _prep(X: pd.DataFrame) -> pd.DataFrame:
        Xp = X.copy().apply(lambda col: col.fillna(col.median(skipna=True)), axis=0)
        return Xp.dropna(axis=1, how="all")

    def _pca2d(X: pd.DataFrame):
        Xp = _prep(X)
        if Xp.shape[0] < 3 or Xp.shape[1] < 2:
            return None, None
        pca = _PCA(n_components=2)
        coords = pca.fit_transform(Xp.values)
        return pd.DataFrame(coords, index=Xp.index, columns=["PC1", "PC2"]), pca.explained_variance_ratio_

    def _scatter(ax, df, labs, title, var):
        for u in sorted(labs.unique()):
            m = labs == u
            ax.scatter(df.loc[m, "PC1"], df.loc[m, "PC2"], label=str(u), alpha=0.7, s=50)
        ax.set_title(title)
        ax.set_xlabel(f"PC1 ({var[0] * 100:.1f}%)")
        ax.set_ylabel(f"PC2 ({var[1] * 100:.1f}%)")

    shared = X_pre.index.intersection(X_post.index)
    if len(shared) < 3:
        logger.warning("PCA skipped — too few shared samples (%d).", len(shared))
        return

    for color_by, labels, fname_suffix in [
        ("batch", batch_labels, "batch"),
        ("group", group_labels, "group"),
    ]:
        labs = labels.reindex(shared).astype(str)
        pre_coords, pre_var   = _pca2d(X_pre.loc[shared])
        post_coords, post_var = _pca2d(X_post.loc[shared])
        if pre_coords is None or post_coords is None:
            logger.warning("PCA (%s) skipped — too few features.", color_by)
            continue

        x_all = np.concatenate([pre_coords["PC1"].values, post_coords["PC1"].values])
        y_all = np.concatenate([pre_coords["PC2"].values, post_coords["PC2"].values])
        xpad  = 0.05 * (x_all.max() - x_all.min() + 1e-9)
        ypad  = 0.05 * (y_all.max() - y_all.min() + 1e-9)

        fig, axes = plt.subplots(1, 2, figsize=(18, 7), sharex=True, sharey=True)
        _scatter(axes[0], pre_coords,  labs,
                 f"PCA pre-ComBat (colored by {color_by})",  pre_var)
        _scatter(axes[1], post_coords, labs,
                 f"PCA post-ComBat (colored by {color_by})", post_var)
        for ax in axes:
            ax.set_xlim(x_all.min() - xpad, x_all.max() + xpad)
            ax.set_ylim(y_all.min() - ypad, y_all.max() + ypad)

        handles, legend_labels = axes[0].get_legend_handles_labels()
        fig.legend(handles, legend_labels, loc="upper right", fontsize=9)
        fig.suptitle(
            f"Lipids Plasma — Pre/Post ComBat PCA (colored by {color_by})", fontsize=14
        )
        fig.tight_layout(rect=[0, 0, 0.97, 0.95])

        out_path = os.path.join(out_dir, f"lipids_pca_pre_post_combat_{fname_suffix}.png")
        fig.savefig(out_path, dpi=200)
        plt.close(fig)
        logger.info("PCA plot saved → %s", out_path)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VALID_TIMEPOINTS = set("ABCDE")

# Columns 0–12 in the processed sheets are feature metadata; data starts at col 13
N_META_COLS = 13

# Row indices in the processed sheets (0-based)
ROW_POLARITY  = 1  # "POS" / "NEG" for each sample column
ROW_SAMPLE_ID = 2  # DP3-XXXXZ or "Pooled Cntrl"
ROW_FILENAME  = 3  # instrument raw filename — used as column header in the sheet
ROW_DATA_START = 4  # ISTD rows and lipid data rows follow


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalise_sample_id(raw_id: str) -> str | None:
    """
    Convert a raw Sample ID label to the canonical DP3-XXXXZ format.

    Rules
    -----
    • Strip internal whitespace between subject number and timepoint letter.
      "DP3-0024 D"  →  "DP3-0024D"
    • Multi-character suffixes after the timepoint (batch replicate labels)
      are stripped; only the first letter of the suffix block is kept.
      "DP3-0258EA"  →  "DP3-0258E"
      "DP3-0136E C" →  "DP3-0136E"
    • Returns None if no valid timepoint can be extracted.
    """
    s = str(raw_id).strip()
    # Normalise any internal whitespace
    s = re.sub(r"\s+", "", s)

    # Match DP3-XXXX followed by one or more letters (timepoint + optional batch suffix)
    m = re.match(r"^(DP3-\d{4})([A-Ea-e].*)$", s)
    if m:
        subject = m.group(1)
        suffix  = m.group(2).upper()
        timepoint = suffix[0]  # keep only the first letter
        if timepoint in VALID_TIMEPOINTS:
            return f"{subject}{timepoint}"
    return None


def _load_lipids_metadata(metadata_path: str) -> pd.DataFrame:
    """
    Load subject-level metadata from the master table (Sheet1).

    Returns a DataFrame indexed by bare DP3-XXXX subject ID with columns:
        Batch, Group, Subgroup, GestAgeDelivery

    Handles postpartum-only subjects whose IDs appear as "DP3-XXXXE" in the
    metadata — the trailing timepoint letter is stripped to create the lookup key,
    so SampleIDs like "DP3-0029E" match correctly after timepoint stripping.
    """
    logger.info("Loading metadata from %s", metadata_path)
    wb = pd.read_excel(metadata_path, sheet_name="Sheet1", header=0, engine="openpyxl")
    rows = []
    n_suffixed = 0
    for _, row in wb.iterrows():
        subject_id = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
        if not subject_id or not subject_id.startswith("DP3-"):
            continue
        gest_age_del = row.iloc[5]  if pd.notna(row.iloc[5])  else None
        group        = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else None
        subgroup     = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else None
        batch        = row.iloc[16] if pd.notna(row.iloc[16]) else None
        if group:
            group = group.replace("sptb", "sPTB")
        # Strip timepoint suffix from postpartum-only subjects ("DP3-0029E" → "DP3-0029")
        m = re.match(r"^(DP3-\d{4})[A-Ea-e]$", subject_id)
        if m:
            lookup_key = m.group(1)
            n_suffixed += 1
        else:
            lookup_key = subject_id
        rows.append({
            "SubjectID": lookup_key,
            "Batch":          batch,
            "Group":          group,
            "Subgroup":       subgroup,
            "GestAgeDelivery": gest_age_del,
        })
    meta_df = pd.DataFrame(rows).set_index("SubjectID")
    meta_df = meta_df[~meta_df.index.duplicated(keep="first")]
    logger.info(
        "Loaded metadata for %d subjects (%d with timepoint suffix in ID column)",
        len(meta_df), n_suffixed,
    )
    return meta_df


def _get_rejected_ids(wb, raw_sheet: str) -> set:
    """
    Returns the set of LipidIDs that are ONLY flagged as Rej=True in the raw
    3Sets sheet (i.e. never appeared as Rej=False).  LipidIDs with at least one
    Rej=False row are considered valid and retained.
    """
    ws = wb[raw_sheet]
    non_rej: set = set()
    rej_only: set = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        rej_flag = row[0]
        lipid_id = row[2]
        if lipid_id is None:
            continue
        if rej_flag is False or rej_flag == 0:
            non_rej.add(lipid_id)
        elif rej_flag is True or rej_flag == 1:
            rej_only.add(lipid_id)
    truly_rejected = rej_only - non_rej
    logger.info(
        "[%s] Non-rejected LipidIDs: %d | Truly-rejected-only: %d",
        raw_sheet, len(non_rej), len(truly_rejected),
    )
    return truly_rejected


def _read_processed_sheet(wb, sheet_name: str, polarity_prefix: str,
                           rejected_ids: set,
                           sid_primary_date: dict | None = None) -> pd.DataFrame:
    """
    Reads a processed sheet ("Plasma POS Lipids" or "Plasma NEG Lipids") and
    returns a DataFrame with:
        rows    = unique canonical SampleIDs (DP3-XXXXA … DP3-XXXXE)
        columns = lipid feature names prefixed with polarity_prefix
        values  = raw peak area (float), NaN for missing

    ISTD rows, Rej=True lipids, and pooled QC columns are excluded.

    Duplicate column handling:
      • Same-batch multi-injections (EA/EB/EC/ED → same canonical SampleID, same
        date batch): averaged together — reduces within-batch technical noise.
      • Cross-batch QC replicates (same canonical SampleID appearing in two
        different date batches): only columns from the sample's primary batch are
        kept; the re-injection column is discarded so that ComBat receives a value
        with a clean, single-batch identity.

    Args:
        sid_primary_date: dict mapping canonical SampleID → primary date batch string
            (e.g. "051223"). If None, no cross-batch filtering is applied (all
            columns are averaged as before).
    """
    import openpyxl  # local import
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    sid_row    = all_rows[ROW_SAMPLE_ID]   # row 2 → sample IDs
    header_row = all_rows[ROW_FILENAME]    # row 3 → filenames / feature meta col names

    # Known LC-MS date batch strings embedded in raw filenames
    _DATE_BATCHES = ("051223", "110123", "041625")

    def _date_batch_of(fname: str) -> str | None:
        for db in _DATE_BATCHES:
            if db in str(fname):
                return db
        return None

    # --- Build column index → (canonical SampleID, date_batch) mapping ---
    col_to_sid: dict[int, str] = {}
    col_to_date: dict[int, str | None] = {}
    for col_idx, (raw_sid, fname) in enumerate(zip(sid_row, header_row)):
        if col_idx < N_META_COLS:
            continue
        if raw_sid is None:
            continue
        raw_sid_str = str(raw_sid).strip()
        if "Pool" in raw_sid_str:
            continue  # skip QC pools
        canon = _normalise_sample_id(raw_sid_str)
        if canon is None:
            logger.debug("Skipping unresolvable SampleID: %r", raw_sid_str)
            continue
        col_to_sid[col_idx] = canon
        col_to_date[col_idx] = _date_batch_of(str(fname))

    # --- Assign accumulator keys: canon_sid__date_batch ---
    # Cross-batch re-injection columns (same SampleID in multiple date batches) are
    # kept as separate accumulator entries so ComBat can use them as reference anchors.
    # Same-batch multi-injection replicates (EA/EB/EC/ED) share the same key and are
    # averaged together before ComBat, as they are within-batch technical replicates.
    # Format: "DP3-0008C__051223" and "DP3-0008C__110123" for cross-batch duplicates.
    col_to_acckey: dict[int, str] = {}
    for col_idx, canon in col_to_sid.items():
        date_batch = col_to_date.get(col_idx)
        if date_batch:
            col_to_acckey[col_idx] = f"{canon}__{date_batch}"
        else:
            col_to_acckey[col_idx] = canon  # fallback: no batch date in filename

    n_cross_batch = sum(
        1 for canon in set(col_to_sid.values())
        if len({col_to_date.get(ci) for ci, c in col_to_sid.items() if c == canon} - {None}) > 1
    )
    if n_cross_batch:
        logger.info(
            "[%s] %d SampleID(s) span multiple date batches — "
            "keeping all for ComBat reference, will average after correction.",
            sheet_name, n_cross_batch,
        )

    patient_cols = sorted(col_to_acckey.keys())
    logger.info(
        "[%s] Patient columns: %d → %d accumulator keys (%d canonical SampleIDs)",
        sheet_name, len(patient_cols),
        len(set(col_to_acckey.values())), len(set(col_to_sid.values())),
    )

    # --- Accumulate peak area data ---
    # accumulator[canonical_sid][feature_name] = list of area values
    accumulator: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))

    n_istd = 0
    n_rejected = 0
    n_kept = 0

    for row_idx in range(ROW_DATA_START, len(all_rows)):
        row = all_rows[row_idx]
        lipid_id    = row[1]
        lipid_group = row[2]

        if lipid_id is None:
            continue

        # Drop ISTD rows (internal standards, not biological analytes)
        if lipid_group is not None and "ISTD" in str(lipid_group):
            n_istd += 1
            continue

        # Drop features rejected by LipidSearch QC in the raw 3Sets sheet
        if lipid_id in rejected_ids:
            n_rejected += 1
            continue

        feature_name = f"{polarity_prefix}__{lipid_id}"
        n_kept += 1

        for col_idx in patient_cols:
            acc_key = col_to_acckey[col_idx]
            area = row[col_idx]
            if area is not None and isinstance(area, (int, float)):
                accumulator[acc_key][feature_name].append(float(area))

    logger.info(
        "[%s] Rows — kept: %d | ISTD dropped: %d | QC-rejected dropped: %d",
        sheet_name, n_kept, n_istd, n_rejected,
    )

    # --- For each (sample, feature): average areas across technical replicates ---
    # (accounts for both multi-batch duplicates and same-compound duplicate rows)
    all_features = set()
    for sid_data in accumulator.values():
        all_features.update(sid_data.keys())
    all_features = sorted(all_features)

    all_sids = sorted(accumulator.keys())

    data_dict: dict[str, list] = {feat: [] for feat in all_features}
    index = []

    for sid in all_sids:
        index.append(sid)
        sid_data = accumulator[sid]
        for feat in all_features:
            vals = sid_data.get(feat, [])
            if vals:
                # Average multiple measurements (technical replicates: same biological
                # sample injected across batches or as EA/EB/EC/ED multi-injections).
                data_dict[feat].append(sum(vals) / len(vals))
            else:
                data_dict[feat].append(np.nan)

    df = pd.DataFrame(data_dict, index=index)
    df.index.name = "SampleID"
    logger.info(
        "[%s] Output matrix: %d rows × %d features (before log2; "
        "cross-batch duplicates kept as separate rows)",
        sheet_name, df.shape[0], df.shape[1],
    )
    return df


def _log2_transform(df: pd.DataFrame) -> pd.DataFrame:
    """Apply log2(x + 1) to all values — +1 pseudo-count handles true zeros."""
    transformed = np.log2(df.values.astype(float) + 1)
    return pd.DataFrame(transformed, index=df.index, columns=df.columns)


def _attach_metadata(df: pd.DataFrame, meta_df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepend SubjectID, Batch, Group, Subgroup, GestAgeDelivery columns to the
    wide lipid matrix. Lookup is done by stripping the timepoint suffix from
    each SampleID to get the bare DP3-XXXX subject key.
    """
    subject_ids    = []
    batches        = []
    groups         = []
    subgroups      = []
    gest_age_dels  = []

    for sid in df.index:
        m = re.match(r"^(DP3-\d{4})[A-E]$", sid)
        subject = m.group(1) if m else None
        if subject and subject in meta_df.index:
            row = meta_df.loc[subject]
            subject_ids.append(subject)
            batches.append(row["Batch"])
            groups.append(row["Group"])
            subgroups.append(row["Subgroup"])
            gest_age_dels.append(row["GestAgeDelivery"])
        else:
            subject_ids.append(None)
            batches.append(None)
            groups.append(None)
            subgroups.append(None)
            gest_age_dels.append(None)

    out = df.copy()
    out.insert(0, "GestAgeDelivery", gest_age_dels)
    out.insert(0, "Subgroup",        subgroups)
    out.insert(0, "Group",           groups)
    out.insert(0, "Batch",           batches)
    out.insert(0, "SubjectID",       subject_ids)
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_cleaned_matrix(
    excel_path: str,
    metadata_path: str,
    output_full_dir: str,
    output_sliced_dir: str,
) -> None:
    import openpyxl

    os.makedirs(output_full_dir, exist_ok=True)
    os.makedirs(output_sliced_dir, exist_ok=True)

    logger.info("Opening workbook: %s", excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # --- Step 1: QC rejection flags from raw 3Sets sheets ---
    logger.info("Reading QC rejection flags from raw 3Sets sheets …")
    rej_pos = _get_rejected_ids(wb, "041625_Sadovsky_Plasma_3Sets_Po")
    rej_neg = _get_rejected_ids(wb, "Sadovsky_Plasma_3Sets_Neg")

    # --- Build primary-batch map: canonical SampleID → date batch string ---
    # Used to discard cross-batch re-injection columns while keeping same-batch
    # multi-injection columns (EA/EB/EC/ED) intact.
    logger.info("Loading metadata to build primary-batch map …")
    meta_df_early = _load_lipids_metadata(metadata_path)
    _OMICS_TO_DATE = {1: "051223", 2: "110123", 3: "041625"}
    sid_primary_date: dict[str, str] = {}
    # We don't know canonical SampleIDs yet; build subject → date, then expand later.
    # Store as subject-level first; _read_processed_sheet will look up by canonical SID.
    _subject_to_date: dict[str, str] = {}
    for subj, row in meta_df_early.iterrows():
        batch_num = row["Batch"]
        if pd.notna(batch_num):
            try:
                date = _OMICS_TO_DATE.get(int(batch_num))
            except (ValueError, TypeError):
                date = None  # e.g. "excluded" or other non-numeric labels
            if date:
                _subject_to_date[subj] = date

    # We need canonical SampleID → date, but those aren't known until the sheet is
    # read. Pass a callable that resolves on demand instead.
    def _get_primary_date_for_sid(canon_sid: str) -> str | None:
        subject = canon_sid[:8]  # "DP3-XXXX"
        return _subject_to_date.get(subject)

    # Pre-build the lookup dict over all possible SampleIDs using known subjects.
    # _read_processed_sheet receives this and filters during column building.
    # We use a defaultdict proxy so it resolves on any canonical SampleID.
    class _PrimaryDateMap:
        def get(self, canon_sid, default=None):
            return _get_primary_date_for_sid(canon_sid) or default

    sid_primary_date_map = _PrimaryDateMap()

    # --- Step 2-3: Read POS and NEG processed sheets → raw area DataFrames ---
    logger.info("Reading POS sheet …")
    df_pos = _read_processed_sheet(wb, "Plasma POS Lipids", "POS", rej_pos,
                                   sid_primary_date=sid_primary_date_map)

    logger.info("Reading NEG sheet …")
    df_neg = _read_processed_sheet(wb, "Plasma NEG Lipids", "NEG", rej_neg,
                                   sid_primary_date=sid_primary_date_map)

    # --- Step 4: Merge POS + NEG (outer join on SampleID) ---
    logger.info("Merging POS and NEG matrices …")
    df_combined = df_pos.join(df_neg, how="outer")
    logger.info(
        "Combined matrix: %d samples × %d features",
        df_combined.shape[0], df_combined.shape[1],
    )

    # --- Step 5: Log2(area + 1) transformation ---
    logger.info("Applying log2(area + 1) transformation …")
    df_log2 = _log2_transform(df_combined)

    # --- Step 6: Use already-loaded metadata ---
    meta_df = meta_df_early

    # Helper: strip batch suffix and timepoint to get bare subject ID.
    # Handles both plain "DP3-0071A" and temp-keyed "DP3-0071A__051223" forms.
    def _subject_of(sid: str) -> str | None:
        sid_clean = sid.split("__")[0]  # strip "__batchdate" if present
        m = re.match(r"^(DP3-\d{4})[A-E]$", sid_clean)
        return m.group(1) if m else None

    # Build original SampleID → temp-key mapping for post-ComBat averaging.
    # If a row index contains "__", it is a cross-batch duplicate temp key.
    has_temp_keys = any("__" in sid for sid in df_log2.index)
    if has_temp_keys:
        original_sid_per_row = pd.Series(
            {sid: sid.split("__")[0] for sid in df_log2.index},
            name="OriginalSID",
        )
        n_dup = (original_sid_per_row.duplicated(keep=False)).sum()
        logger.info(
            "Cross-batch duplicate rows detected: %d rows will be averaged "
            "post-ComBat to yield unique SampleIDs.", n_dup,
        )
    else:
        original_sid_per_row = None

    # Per-sample group labels (from metadata, using original SampleID)
    groups = pd.Series(
        {sid: (meta_df.loc[_subject_of(sid), "Group"]
               if _subject_of(sid) and _subject_of(sid) in meta_df.index
               else None)
         for sid in df_log2.index},
        name="Group",
    )

    # ComBat batch labels: use the date-batch encoded in the temp key (most accurate).
    # For rows without a temp key, fall back to metadata batch converted to date string.
    _OMICS_TO_DATE_INV = {"051223": "051223", "110123": "110123", "041625": "041625"}
    def _combat_batch(sid: str) -> str | None:
        if "__" in sid:
            return sid.split("__")[1]  # date batch from temp key
        # Fallback: derive from metadata omics set number
        subj = _subject_of(sid)
        if subj and subj in meta_df.index:
            batch_num = meta_df.loc[subj, "Batch"]
            if pd.notna(batch_num):
                return {1: "051223", 2: "110123", 3: "041625"}.get(int(batch_num))
        return None

    batch_labels = pd.Series(
        {sid: _combat_batch(sid) for sid in df_log2.index},
        name="Batch",
    )

    # Metadata matching diagnostics
    n_matched = groups.notna().sum()
    n_total = len(groups)
    if n_matched < n_total:
        unmatched = groups[groups.isna()].index.tolist()
        logger.warning(
            "%d samples missing metadata — examples: %s",
            n_total - n_matched, unmatched[:5],
        )
    logger.info("Metadata matched: %d / %d samples", n_matched, n_total)

    # --- Step 7a: Filter samples missing group (needed before ComBat) ---
    has_metadata = groups.notna()
    n_no_meta = int((~has_metadata).sum())
    if n_no_meta > 0:
        logger.warning("Filtering %d samples without metadata group.", n_no_meta)
        df_log2      = df_log2.loc[has_metadata].copy()
        groups       = groups.loc[has_metadata]
        batch_labels = batch_labels.loc[has_metadata]

    # Build binary group labels (Control vs Complication) for missingness test
    groups_binary = pd.Series(
        np.where(
            groups.astype(str).str.strip().str.upper() == "CONTROL",
            "Control",
            "Complication",
        ),
        index=groups.index,
        name="GroupBinary",
    )

    # --- Step 7b: Filter samples missing batch labels ---
    has_batch = batch_labels.notna()
    n_no_batch = int((~has_batch).sum())
    if n_no_batch > 0:
        logger.warning("Filtering %d samples without batch labels.", n_no_batch)
        df_log2       = df_log2.loc[has_batch].copy()
        groups_binary = groups_binary.loc[has_batch].copy()
        batch_labels  = batch_labels.loc[has_batch]

    # --- Step 8: ComBat batch correction ---
    logger.info(
        "Applying ComBat batch correction (batches: %s) …",
        sorted(batch_labels.unique()),
    )
    X_norm = combat_normalize_wide(df_log2, batch_labels)

    # --- Step 8b: Pre/post ComBat PCA plots ---
    logger.info("Generating pre/post ComBat PCA plots …")
    _generate_pca_plots(
        X_pre=df_log2,
        X_post=X_norm,
        batch_labels=batch_labels.astype(str),
        group_labels=groups,
        out_dir=output_full_dir,
    )

    # --- Step 8c: Average cross-batch duplicates post-ComBat ---
    if original_sid_per_row is not None:
        orig = original_sid_per_row.reindex(X_norm.index)
        n_before = len(X_norm)
        X_norm        = X_norm.groupby(orig).mean()
        groups        = groups.groupby(orig).first()
        batch_labels  = batch_labels.groupby(orig).first()
        groups_binary = groups_binary.groupby(orig).first()
        logger.info(
            "Post-ComBat averaging: %d rows → %d unique SampleIDs.",
            n_before, len(X_norm),
        )

    # --- Step 9: Missingness filter (≥20% → drop) + Fisher/BH report ---
    logger.info(
        "Applying missingness filter (cutoff = %.0f%%) …",
        CUTOFF_PERCENT_MISSING * 100,
    )
    X_kept, dropped_report = missingness_filter_and_group_check(
        X_norm, groups_binary, cutoff=CUTOFF_PERCENT_MISSING, alpha_bh=0.05,
    )
    n_dropped = len(dropped_report)
    if n_dropped > 0:
        rep_path = os.path.join(
            output_full_dir, "lipids_plasma_dropped_missingness_report.csv"
        )
        dropped_report.to_csv(rep_path, index=False)
        logger.info(
            "Dropped %d features (missingness ≥ 20%%) — report → %s",
            n_dropped, rep_path,
        )
    logger.info(
        "Features after missingness filter: %d (dropped %d)",
        X_kept.shape[1], n_dropped,
    )

    # --- Step 10: Half-minimum imputation (log2 space: fill = min_log2 − 1) ---
    n_missing = int(X_kept.isna().sum().sum())
    logger.info("Missing values before imputation: %d", n_missing)
    X_final = half_min_impute_wide(X_kept)
    logger.info("Imputation complete.")

    # --- Step 11: Attach metadata columns ---
    logger.info("Attaching metadata columns …")
    df_final = _attach_metadata(X_final, meta_df)

    # Report samples without metadata (should be 0 after the filter above)
    no_meta = df_final[df_final["Group"].isna()].index.tolist()
    if no_meta:
        logger.warning(
            "%d samples had no metadata match: %s",
            len(no_meta), no_meta[:10],
        )

    # --- Step 12: Save full matrix ---
    full_csv = os.path.join(output_full_dir, "lipids_plasma_cleaned_with_metadata.csv")
    df_final.to_csv(full_csv)
    logger.info(
        "Saved full matrix → %s  (%d rows × %d cols)",
        full_csv, *df_final.shape,
    )

    # --- Step 13: Slice by timepoint suffix (A–E) and save ---
    logger.info("Slicing by timepoint suffix …")
    for tp in "ABCDE":
        mask = df_final.index.str.endswith(tp)
        df_tp = df_final[mask].copy()

        # Remove the timepoint suffix from SampleID to match pipeline convention
        df_tp.index = df_tp.index.str[:-1]   # "DP3-0217A" → "DP3-0217"
        df_tp.index.name = "SampleID"

        out_path = os.path.join(
            output_sliced_dir, f"lipids_plasma_formatted_suffix_{tp}.csv"
        )
        df_tp.to_csv(out_path)
        logger.info(
            "  Timepoint %s: %d samples → %s", tp, len(df_tp), out_path
        )

    logger.info(
        "Done. Final matrix: %d samples × %d lipid features | dropped features: %d",
        X_final.shape[0], X_final.shape[1], n_dropped,
    )


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Format Sadovsky plasma lipidomics data for the DP3 pipeline."
    )
    p.add_argument(
        "--excel",
        default=os.path.join(
            "data", "lipids", "072925 Sadovsky Plasma Lipids Untargeted ALL.xlsx"
        ),
        help="Path to the lipidomics Excel workbook.",
    )
    p.add_argument(
        "--metadata",
        default=os.path.join("data", "dp3 master table v2.xlsx"),
        help="Path to DP3 master metadata Excel file.",
    )
    p.add_argument(
        "--output-full-dir",
        default=os.path.join("data", "cleaned", "lipids", "normalized_full_results"),
        help="Directory for the full merged matrix CSV.",
    )
    p.add_argument(
        "--output-sliced-dir",
        default=os.path.join(
            "data", "cleaned", "lipids", "normalized_sliced_by_suffix"
        ),
        help="Directory for per-timepoint CSV slices.",
    )
    return p


if __name__ == "__main__":
    args = _build_parser().parse_args()
    build_cleaned_matrix(
        excel_path=args.excel,
        metadata_path=args.metadata,
        output_full_dir=args.output_full_dir,
        output_sliced_dir=args.output_sliced_dir,
    )
